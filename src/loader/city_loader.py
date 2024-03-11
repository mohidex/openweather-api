import io
import aiohttp
from typing import AsyncIterable
from collections import namedtuple
from openpyxl import load_workbook

CityData = namedtuple("CityData", ['name', 'lon', 'lat', 'country'])


class CityLoader:
    """
    Asynchronous loader for city data from an Excel file.
    attributes:
        url (str): The URL of the Excel file to download.
        override (bool): Whether to override the existing data in the database.
    """

    def __init__(self, url: str, override: bool = False):
        self.url = url
        self.override = override

    async def download_excel_file(self) -> io.BytesIO:
        async with aiohttp.ClientSession(raise_for_status=True) as http_client:
            async with http_client.get(self.url) as response:
                content = await response.read()
                return io.BytesIO(content)

    async def get_cities(self) -> AsyncIterable[CityData]:
        excel_file = await self.download_excel_file()
        workbook = load_workbook(filename=excel_file, read_only=True)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            yield CityData(*row)
        workbook.close()
        excel_file.close()
